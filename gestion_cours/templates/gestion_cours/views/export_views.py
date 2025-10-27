from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
# Remonter d'un niveau pour importer les modèles (depuis gestion_cours)
from ..models import MatiereProgrammee  # type: ignore




# Le décorateur login_required assure que seul un utilisateur connecté peut accéder à cette vue.
@login_required
def export_cours_to_excel(request):
    """
    Exporte la liste des MatiereProgrammee vers un fichier Excel (format .xlsx).
    """
    
    # 1. Préparation de la réponse HTTP pour le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Nom du fichier avec horodatage
    filename = f"Cours_Programmes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 2. Création du classeur et de la feuille de calcul
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cours Programmés"

    # 3. Définition des en-têtes et des largeurs de colonnes
    columns = [
        ('Matière', 25),
        ('Filière', 15),
        ('Niveau', 10),
        ('Professeur', 30),
        ('Année Académique', 20),
        ('Volume Horaire Prévu', 20),
        ('Date Début Estimée', 20),
        ('Date Fin Estimée', 20),
        ('Semestre', 10),
    ]

    row_num = 1
    # Écriture des en-têtes
    for col_num, (header, width) in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}{row_num}'] = header
        worksheet.column_dimensions[col_letter].width = width

    # 4. Récupération des données (Optimisé pour éviter les requêtes N+1)
    cours = MatiereProgrammee.objects.select_related(
        'matiere', 'filiere', 'niveau', 'professeur', 'annee_academique'
    ).all()

    # 5. Remplissage des données
    for cours_obj in cours:
        row_num += 1
        
        # Formatage du nom du niveau
        niveau_display = f"{cours_obj.niveau.libelle} {cours_obj.niveau.niv}" if cours_obj.niveau else ""
        
        # Formatage du nom du professeur
        prof_display = f"{cours_obj.professeur.nom} {cours_obj.professeur.prenoms}" if cours_obj.professeur else ""
        
        # Conversion du Decimal en float pour assurer la compatibilité Excel
        volume_horaire = float(cours_obj.nbr_heure) if cours_obj.nbr_heure else 0.0

        worksheet.cell(row=row_num, column=1, value=cours_obj.matiere.libelle)
        worksheet.cell(row=row_num, column=2, value=cours_obj.filiere.libelle)
        worksheet.cell(row=row_num, column=3, value=niveau_display)
        worksheet.cell(row=row_num, column=4, value=prof_display)
        worksheet.cell(row=row_num, column=5, value=cours_obj.annee_academique.annee_accademique)
        worksheet.cell(row=row_num, column=6, value=volume_horaire)
        worksheet.cell(row=row_num, column=7, value=cours_obj.date_debut_estimee)
        worksheet.cell(row=row_num, column=8, value=cours_obj.date_fin_estimee)
        worksheet.cell(row=row_num, column=9, value=cours_obj.semestre)

    # 6. Sauvegarde du classeur dans la réponse HTTP
    workbook.save(response)
    return response
