from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime


# ======================== STYLES AMÉLIORÉS ========================

def get_excel_styles():
    """
    Retourne des styles modernes et élégants pour Excel.
    Palette de couleurs professionnelle : Bleu marine et or
    """
    # Palette de couleurs moderne
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Bleu marine profond
    header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Bordures élégantes
    thin_border = Border(
        left=Side(style='thin', color='B4C7E7'),
        right=Side(style='thin', color='B4C7E7'),
        top=Side(style='thin', color='B4C7E7'),
        bottom=Side(style='thin', color='B4C7E7')
    )
    
    # Style pour les cellules de données
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_font = Font(size=11, name='Calibri', color='1F2937')
    
    # Alternance de couleurs pour les lignes
    row_fill_even = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'header_alignment': header_alignment,
        'border': thin_border,
        'data_alignment': data_alignment,
        'data_font': data_font,
        'row_fill_even': row_fill_even,
        'row_fill_odd': row_fill_odd
    }


def apply_excel_styling(ws, styles, num_columns, has_data=True):
    """
    Applique un style cohérent à toute la feuille Excel.
    """
    # Style de l'en-tête (ligne 1)
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    if has_data:
        # Style des données avec alternance de couleurs
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.font = styles['data_font']
                cell.alignment = styles['data_alignment']
                cell.border = styles['border']
                # Alternance de couleurs
                if row_idx % 2 == 0:
                    cell.fill = styles['row_fill_even']
                else:
                    cell.fill = styles['row_fill_odd']
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'


def setup_pdf_styles():
    """Configure des styles PDF modernes avec support des accents."""
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        font_name = 'DejaVu'
        font_name_bold = 'DejaVu-Bold'
    except:
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
    
    styles = getSampleStyleSheet()
    
    # Titre principal - Design moderne
    styles.add(ParagraphStyle(
        name='ModernTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1E3A8A'),  # Bleu marine
        spaceAfter=20,
        spaceBefore=10,
        alignment=1,
        fontName=font_name_bold,
        leading=24
    ))
    
    # Sous-titre
    styles.add(ParagraphStyle(
        name='Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#6B7280'),  # Gris
        spaceAfter=25,
        alignment=1,
        fontName=font_name,
        leading=14
    ))
    
    # Note de bas de page
    styles.add(ParagraphStyle(
        name='Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=0,
        fontName=font_name,
        leading=10
    ))
    
    return styles, font_name, font_name_bold


def get_pdf_table_style(font_name):
    """Retourne un style de tableau PDF moderne et élégant."""
    return TableStyle([
        # En-tête - Bleu marine élégant
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, 0), 14),
        
        # Données
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1F2937')),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 1), (-1, -1), 8),
        ('RIGHTPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        
        # Bordures subtiles
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1E3A8A')),
        ('LINEBELOW', (0, 1), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        
        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ])


# ======================== EXPORT PROFESSEURS ========================

def export_professeurs_to_excel(professeurs_list):
    """Exporte la liste des professeurs vers Excel avec un design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Professeurs"
    
    styles = get_excel_styles()
    
    # En-tête
    headers = ['Nom', 'Prénoms', 'Contact', 'Email', 'Grade', "Domaine d'expertise"]
    ws.append(headers)
    
    # Données
    for prof in professeurs_list:
        ws.append([
            prof.nom,
            prof.prenoms,
            prof.contact,
            prof.email,
            prof.grade or 'N/A',
            prof.domaine or 'N/A'
        ])
    
    # Appliquer les styles
    apply_excel_styling(ws, styles, len(headers), has_data=len(professeurs_list) > 0)
    
    # Ajuster les largeurs
    column_widths = [18, 20, 15, 25, 18, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Hauteur de l'en-tête
    ws.row_dimensions[1].height = 35
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Professeurs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_professeurs_to_pdf(professeurs_list):
    """Exporte la liste des professeurs vers PDF avec un design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    # Titre
    title = Paragraph("Liste des Professeurs", styles['ModernTitle'])
    elements.append(title)
    
    # Sous-titre avec date
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(professeurs_list)} professeur(s)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    # Données du tableau
    data = [['Nom', 'Prénoms', 'Contact', 'Email', 'Grade', "Domaine"]]
    
    for prof in professeurs_list:
        data.append([
            Paragraph(prof.nom, styles['Normal']),
            Paragraph(prof.prenoms, styles['Normal']),
            prof.contact,
            Paragraph(prof.email, styles['Normal']),
            prof.grade or 'N/A',
            Paragraph(prof.domaine or 'N/A', styles['Normal'])
        ])
    
    # Créer le tableau
    table = Table(data, colWidths=[3.5*cm, 3.5*cm, 3*cm, 5*cm, 3*cm, 5.5*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph(
        "Document confidentiel - Usage interne uniquement",
        styles['Footer']
    )
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Professeurs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT COURS PROGRAMMÉS ========================

def export_cours_to_excel(cours_list):
    """Exporte la liste des cours programmés vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cours Programmés"
    
    styles = get_excel_styles()
    
    # En-tête
    headers = ['Matière', 'Filière', 'Niveau', 'Professeur', 'Semestre', 
               'Volume Horaire', 'Heures Faites', 'Progression', 'Date Début', 'Date Fin']
    ws.append(headers)
    
    # Données
    for cours in cours_list:
        heures_faites = cours.total_heures_faites if hasattr(cours, 'total_heures_faites') else 0
        volume = float(cours.nbr_heure)
        progression = (float(heures_faites or 0) / volume * 100) if volume > 0 else 0
        
        ws.append([
            cours.matiere.libelle,
            cours.filiere.code,
            f"{cours.niveau.libelle} {cours.niveau.niv}",
            str(cours.professeur),
            cours.get_semestre_display(),
            volume,
            float(heures_faites or 0),
            f"{progression:.1f}%",
            cours.date_debut_estimee.strftime('%d/%m/%Y') if cours.date_debut_estimee else 'N/A',
            cours.date_fin_estimee.strftime('%d/%m/%Y') if cours.date_fin_estimee else 'N/A'
        ])
    
    # Appliquer les styles
    apply_excel_styling(ws, styles, len(headers), has_data=len(cours_list) > 0)
    
    # Ajuster les largeurs
    column_widths = [25, 12, 15, 22, 12, 14, 14, 12, 14, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Hauteur de l'en-tête
    ws.row_dimensions[1].height = 40
    
    # Mise en forme conditionnelle pour la progression
    for row in range(2, len(cours_list) + 2):
        cell = ws.cell(row=row, column=8)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Couleur selon progression
        progression_val = float(cell.value.strip('%'))
        if progression_val >= 100:
            cell.font = Font(bold=True, color='10B981', size=11)  # Vert
        elif progression_val >= 75:
            cell.font = Font(bold=True, color='3B82F6', size=11)  # Bleu
        elif progression_val >= 50:
            cell.font = Font(bold=True, color='F59E0B', size=11)  # Orange
        else:
            cell.font = Font(bold=True, color='EF4444', size=11)  # Rouge
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Cours_Programmes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_cours_to_pdf(cours_list):
    """Exporte la liste des cours programmés vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    # Titre
    title = Paragraph("Cours Programmés", styles['ModernTitle'])
    elements.append(title)
    
    # Sous-titre
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(cours_list)} cours",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    # Données du tableau
    data = [['Matière', 'Filière', 'Niveau', 'Professeur', 'Sem.', 'Vol. H', 'H. Faites', 'Prog.']]
    
    for cours in cours_list:
        heures_faites = cours.total_heures_faites if hasattr(cours, 'total_heures_faites') else 0
        volume = float(cours.nbr_heure)
        progression = (float(heures_faites or 0) / volume * 100) if volume > 0 else 0
        
        data.append([
            Paragraph(cours.matiere.libelle[:20], styles['Normal']),
            cours.filiere.code,
            f"{cours.niveau.libelle} {cours.niveau.niv}",
            Paragraph(str(cours.professeur)[:20], styles['Normal']),
            cours.get_semestre_display(),
            f"{volume}h",
            f"{float(heures_faites or 0)}h",
            f"{progression:.0f}%"
        ])
    
    # Créer le tableau
    table = Table(data, colWidths=[4.5*cm, 2*cm, 2.5*cm, 4*cm, 1.8*cm, 1.8*cm, 2*cm, 1.8*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph("Document confidentiel - Usage interne uniquement", styles['Footer'])
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Cours_Programmes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT ÉMARGEMENTS ========================

def export_emargements_to_excel(emargements_list):
    """Exporte la liste des émargements vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Émargements"
    
    styles = get_excel_styles()
    
    # En-tête
    headers = ['Date', 'Matière', 'Professeur', 'Filière', 'Niveau', 'Semestre', 'Heures Effectuées']
    ws.append(headers)
    
    # Données
    for emarg in emargements_list:
        ws.append([
            emarg.date_emar.strftime('%d/%m/%Y'),
            emarg.matiere_programmer.matiere.libelle,
            str(emarg.matiere_programmer.professeur),
            emarg.matiere_programmer.filiere.code,
            f"{emarg.matiere_programmer.niveau.libelle} {emarg.matiere_programmer.niveau.niv}",
            emarg.matiere_programmer.get_semestre_display(),
            float(emarg.heure_eff)
        ])
    
    # Appliquer les styles
    apply_excel_styling(ws, styles, len(headers), has_data=len(emargements_list) > 0)
    
    # Ajuster les largeurs
    column_widths = [14, 28, 25, 12, 18, 12, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 35
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Emargements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_emargements_to_pdf(emargements_list):
    """Exporte la liste des émargements vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    title = Paragraph("Registre des Émargements", styles['ModernTitle'])
    elements.append(title)
    
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(emargements_list)} émargement(s)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    data = [['Date', 'Matière', 'Professeur', 'Filière', 'Niveau', 'Sem.', 'Heures']]
    
    for emarg in emargements_list:
        data.append([
            emarg.date_emar.strftime('%d/%m/%Y'),
            Paragraph(emarg.matiere_programmer.matiere.libelle[:20], styles['Normal']),
            Paragraph(str(emarg.matiere_programmer.professeur)[:20], styles['Normal']),
            emarg.matiere_programmer.filiere.code,
            f"{emarg.matiere_programmer.niveau.libelle} {emarg.matiere_programmer.niveau.niv}",
            emarg.matiere_programmer.get_semestre_display(),
            f"{float(emarg.heure_eff)}h"
        ])
    
    table = Table(data, colWidths=[2.5*cm, 4.5*cm, 4.5*cm, 2*cm, 3*cm, 1.8*cm, 2*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph("Document confidentiel - Usage interne uniquement", styles['Footer'])
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Emargements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT ÉVALUATIONS ========================

def export_evaluations_to_excel(queryset):
    """Exporte les évaluations vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Évaluations"
    
    styles = get_excel_styles()
    
    # En-tête
    headers = [
        'Cours', 
        'Code', 
        'Professeur', 
        'Niveau', 
        'Filière',
        'Résumé Évaluation',
        'Appréciation AP',
        'Recommandations',
        'Évalué par'
    ]
    ws.append(headers)
    
    # Données
    for eval_obj in queryset:
        eval_par = eval_obj.utilisateur_evaluation.get_full_name() if eval_obj.utilisateur_evaluation and hasattr(eval_obj.utilisateur_evaluation, 'get_full_name') else (eval_obj.utilisateur_evaluation.username if eval_obj.utilisateur_evaluation else 'Système')
        
        ws.append([
            eval_obj.matiere_programmer.matiere.libelle,
            eval_obj.matiere_programmer.matiere.code,
            str(eval_obj.matiere_programmer.professeur),
            f"{eval_obj.matiere_programmer.niveau.libelle} {eval_obj.matiere_programmer.niveau.niv}", 
            eval_obj.matiere_programmer.filiere.code, 
            eval_obj.resume_evaluation or 'Non renseigné',
            eval_obj.resume_ap or 'Non renseigné',
            eval_obj.recommandation or 'Non renseigné',
            eval_par
        ])
    
    # Appliquer les styles
    apply_excel_styling(ws, styles, len(headers), has_data=len(queryset) > 0)
    
    # Ajuster les largeurs
    column_widths = [28, 12, 22, 18, 12, 35, 35, 35, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 40
    
    # Hauteur automatique pour les lignes de données
    for row in range(2, len(queryset) + 2):
        ws.row_dimensions[row].height = None  # Auto
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Evaluations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_evaluations_to_pdf(queryset):
    """Exporte les évaluations vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    title = Paragraph("Évaluations Qualitatives des Enseignements", styles['ModernTitle'])
    elements.append(title)
    
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(queryset)} évaluation(s)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    # Données du tableau
    data = [['Cours', 'Professeur', 'Résumé Évaluation', 'Appréciation AP', 'Recommandations', 'Évalué par']]
    
    table_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=7,
        fontName=font_name,
        alignment=0,
        leading=9,
    )
    
    for eval_obj in queryset:
        resume = eval_obj.resume_evaluation[:120] + '...' if eval_obj.resume_evaluation and len(eval_obj.resume_evaluation) > 120 else (eval_obj.resume_evaluation or 'N/A')
        appreciation = eval_obj.resume_ap[:120] + '...' if eval_obj.resume_ap and len(eval_obj.resume_ap) > 120 else (eval_obj.resume_ap or 'N/A')
        recommandations_txt = eval_obj.recommandation[:120] + '...' if eval_obj.recommandation and len(eval_obj.recommandation) > 120 else (eval_obj.recommandation or 'N/A')
        eval_par = eval_obj.utilisateur_evaluation.get_full_name() if eval_obj.utilisateur_evaluation and hasattr(eval_obj.utilisateur_evaluation, 'get_full_name') else (eval_obj.utilisateur_evaluation.username if eval_obj.utilisateur_evaluation else 'Système')

        data.append([
            Paragraph(eval_obj.matiere_programmer.matiere.libelle, table_style),
            Paragraph(str(eval_obj.matiere_programmer.professeur), table_style),
            Paragraph(resume, table_style),
            Paragraph(appreciation, table_style),
            Paragraph(recommandations_txt, table_style),
            Paragraph(eval_par, table_style)
        ])
    
    table = Table(data, colWidths=[4*cm, 3.5*cm, 5*cm, 5*cm, 5*cm, 3*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    elements.append(Spacer(1, 0.8*cm))
    footer = Paragraph(
        "<b>Note :</b> Les textes dépassant 120 caractères sont tronqués. Consultez l'interface web ou l'export Excel pour le contenu complet.",
        styles['Footer']
    )
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Evaluations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT MATIÈRES ========================

def export_matieres_to_excel(matieres_list):
    """Exporte la liste des matières vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Matières"
    
    styles = get_excel_styles()
    
    headers = ['Code', 'Libellé']
    ws.append(headers)
    
    for matiere in matieres_list:
        ws.append([matiere.code, matiere.libelle])
    
    apply_excel_styling(ws, styles, len(headers), has_data=len(matieres_list) > 0)
    
    column_widths = [18, 45]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 35
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Matieres_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_matieres_to_pdf(matieres_list):
    """Exporte la liste des matières vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    title = Paragraph("Catalogue des Matières", styles['ModernTitle'])
    elements.append(title)
    
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(matieres_list)} matière(s)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    data = [['Code', 'Libellé']]
    
    for matiere in matieres_list:
        data.append([
            matiere.code,
            Paragraph(matiere.libelle, styles['Normal'])
        ])
    
    table = Table(data, colWidths=[4*cm, 12*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph("Document confidentiel - Usage interne uniquement", styles['Footer'])
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Matieres_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT FILIÈRES ========================

def export_filieres_to_excel(filieres_list):
    """Exporte la liste des filières vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Filières"
    
    styles = get_excel_styles()
    
    headers = ['Code', 'Libellé']
    ws.append(headers)
    
    for filiere in filieres_list:
        ws.append([filiere.code, filiere.libelle])
    
    apply_excel_styling(ws, styles, len(headers), has_data=len(filieres_list) > 0)
    
    column_widths = [15, 45]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 35
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Filieres_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_filieres_to_pdf(filieres_list):
    """Exporte la liste des filières vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    title = Paragraph("Catalogue des Filières", styles['ModernTitle'])
    elements.append(title)
    
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(filieres_list)} filière(s)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    data = [['Code', 'Libellé']]
    
    for filiere in filieres_list:
        data.append([
            filiere.code,
            Paragraph(filiere.libelle, styles['Normal'])
        ])
    
    table = Table(data, colWidths=[3.5*cm, 12*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph("Document confidentiel - Usage interne uniquement", styles['Footer'])
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Filieres_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ======================== EXPORT NIVEAUX ========================

def export_niveaux_to_excel(niveaux_list):
    """Exporte la liste des niveaux vers Excel avec design moderne."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Niveaux"
    
    styles = get_excel_styles()
    
    headers = ['Niveau', 'Libellé', 'Filière', 'Code Filière']
    ws.append(headers)
    
    for niveau in niveaux_list:
        ws.append([
            niveau.niv,
            niveau.libelle,
            niveau.filiere.libelle,
            niveau.filiere.code
        ])
    
    apply_excel_styling(ws, styles, len(headers), has_data=len(niveaux_list) > 0)
    
    column_widths = [12, 25, 30, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 35
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Niveaux_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def export_niveaux_to_pdf(niveaux_list):
    """Exporte la liste des niveaux vers PDF avec design moderne."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    styles, font_name, font_name_bold = setup_pdf_styles()
    
    title = Paragraph("Catalogue des Niveaux", styles['ModernTitle'])
    elements.append(title)
    
    subtitle = Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {len(niveaux_list)} niveau(x)",
        styles['Subtitle']
    )
    elements.append(subtitle)
    
    data = [['Niveau', 'Libellé', 'Filière', 'Code']]
    
    for niveau in niveaux_list:
        data.append([
            niveau.niv,
            niveau.libelle,
            Paragraph(niveau.filiere.libelle, styles['Normal']),
            niveau.filiere.code
        ])
    
    table = Table(data, colWidths=[2.5*cm, 5*cm, 6*cm, 2.5*cm])
    table.setStyle(get_pdf_table_style(font_name))
    
    elements.append(table)
    
    elements.append(Spacer(1, 1*cm))
    footer = Paragraph("Document confidentiel - Usage interne uniquement", styles['Footer'])
    elements.append(footer)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"Niveaux_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
